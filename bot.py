
import os
import sys
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import glob
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

path = '//10.7.66.75/e/next 25/Chandan Tubes/'
#path = "//10.7.66.75/e$/next 25/Chandan Tubes/"
ReportFilename = "Inspection Report format -Ratnamani.xlsm"

ASMEFilename = "ASME DATA SHEET_Final.xlsx"
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

#ReportOutPutFilename = "2.xlsx"
import openpyxl
import zipfile
from shutil import copyfile
from shutil import rmtree
import pandas as pd
from openpyxl.styles import PatternFill
import shutil
import numpy as np

PAD = os.getcwd()

wb = openpyxl.load_workbook(path+ReportFilename,read_only=False, keep_vba=True)

import pyodbc 

#server = "VHZPDSQLDEV\VHZDVSQL2K16"
#database="Instoman"
#user="Instoman" 
#password="Instoman@123"


server = "PHZPDSQL2K16\PHZPDSQL2K16"
database="MISS_APP"
user="MISS_APP_ADMIN" 
password="MISS_APP_ADMIN@123"

conn =pyodbc.connect(driver='{SQL Server Native Client 11.0}', host=server, database=database, user=user, password=password)

cursor = conn.cursor()

#row = cursor.execute('select * from [instoman1_document]')

#for i in row.fetchall():
#    print(i)


# load final report format
redFill = PatternFill(start_color='FFFF0000',
                end_color='FFFF0000',
                fill_type='solid')
greenFill = PatternFill(start_color = '228B22', end_color='228B22', fill_type='solid')

#l = ['22005ASU1_R0.pdf' , 'PO - 04L013727.PDF', '21921AST1_RO.pdf', 'PO - 04L013842.PDF']
#for i in l:
#    shutil.copy2('//10.7.66.75/e/next 25/Chandan Tubes/' + i , r'//10.7.66.75/e/next 25/Chandan Tubes/Import_documents/'+ i) 

#####
## do magic with openpyxl here and save
#ws = wb.worksheets[1]
#ws.cell(row=5,column=2).value = "Clear"
#print("final report - ",ws.cell(row=5, column=2).value)   # example
#######
def Acceptance(ws,comparestring,tcstring,i):
    print("Additonal function is called")
    ws.cell(row=15+i*2,column=6).value = tcstring
    print(comparestring,tcstring,i)
    if(tcstring.count(comparestring) >0):

        ws.cell(row=15+i*2,column=7).value = "Accepted"
        ws.cell(row=15+i*2,column=7).fill = greenFill 
        ws.cell(row=15+i*2,column=7).border = thin_border
    else:
        ws.cell(row=15+i*2,column=7).value = "Rejected"
        ws.cell(row=15+i*2,column=7).fill = redFill
 #ARM data into asme sheet
def write_arm(wb,CertificationDetailsPath,ARMFilePath,ARMNo,Material1,Grade, Chemical_deviation, Mechanical_deviation,stHeatTreatment,certification,Process,Product,Delivery,list_product):
    # search ARM function
        #ARMFilePath = r'\\10.23.67.120\c$\Users\90347908\Desktop\next 25\TestBatch6\22005ASU1_R0.xlsx'
        #ARMFilePath = r"\\10.7.66.75\e$\next 25\Chandan Tubes\ARM-EXport\TestBatch3\Data_00000001.xlsx"
        print(wb,CertificationDetailsPath,ARMFilePath,ARMNo,Material1,Grade, Chemical_deviation, Mechanical_deviation,stHeatTreatment,certification,Process,Product,Delivery)
        print('ARM comparison started..')


        HeatTreatment_fileName = CertificationDetailsPath +  '\HeatTreatment.xlsx'
        HeatTreatment = pd.read_excel(HeatTreatment_fileName)

        ContentHeatTreatment = ''

        for i in HeatTreatment['Conditions']:
            ContentHeatTreatment+=i

        Remarks_filename = CertificationDetailsPath + '\Remarks.xlsx'
        Remarks = pd.read_excel(Remarks_filename)
        print(Remarks)

        d = {}
        tpi = ''
        Additonal_remarks = []
        last_remark = ''
        for x in Remarks['Remarks']:
            try:
                print(isinstance(int(x.split(' ')[0]),int))
                i = x.lower()
                if(i.count('hydrostatic test') > 0):
                    d['Hydrostatic Test'] = x
                    last_remark='Hydrostatic Test'
                elif(i.count('igc') > 0):
                    d['IGC'] = x
                    last_remark='IGC'
                    if(len(x.split('by')[-1])>0 and tpi ==''):
                        tpi = x.split('by')[-1]

                elif(i.count('pmi') > 0):
                    d['PMI'] = x
                    last_remark='PMI'
                    if(tpi =='' and len(x.split('by')[-1])>0 ):
                        tpi = x.split('by')[-1]

                elif(i.count('pt examination') > 0):
                    d['PT examination'] = x
                    last_remark='PT examination'
                elif(i.count('visual') > 0):
                    d['Visual'] = x

                elif(i.count('eddy current') > 0 and i.count('heat analysis') < 1  ):
                    d['Eddy Current'] = x
                    last_remark='Eddy Current'
                elif((i.count('product analysis') > 0 or i.count('product chemical analysis') > 0) and i.count('heat analysis') < 1  ):
                    d['Product Analysis'] = x
                    last_remark='Product Analysis'
                elif(i.count('weld') > 0):
                    d['Weld'] = x
                else:
                    Additonal_remarks.append(x)
            except:
                print(last_remark)
                #d[last_remark] = d[last_remark] + x
                pass
        print(d)
        
        Total_deviation = set()
        print(Chemical_deviation,Mechanical_deviation)
        arm = pd.read_excel(ARMFilePath+'.xlsx')

        armdata = arm[["ROW_INDEX","S_NO", "Parameter" , "Reference_Code_Value"]]

        ws = wb.worksheets[0]

        for i in range(len(armdata["ROW_INDEX"])):
            try:

                if(armdata["Parameter"][i].upper().count('PACKAGING') > 0):
                    break
                
                if(armdata["S_NO"][i].upper().count('LARSE') > 0 or armdata["S_NO"][i].upper().count('LARS') > 0):
                    continue

                
                if( str(armdata["S_NO"][i]) == 'B1' or str(armdata["S_NO"][i]) == 'B2'):
                    print('basic and additional chemistry')
                    continue

                    #ws.cell(row=15+i*2,column=3).value = 'Basic Chemistry Requirements'
 
                    #ws.cell(row=15+i*2,column=3).border = thin_border

                    
                    #ws.cell(row=15+i*2,column=8).value = "Chemistry is reported in Chemical sheet"
                    #ws.cell(row=15+i*2,column=8).border = thin_border

                    #continue
                
                #if( armdata["S_NO"][i] == 'B2'):

                #    ws.cell(row=15+i*2,column=3).value = 'Additional Chemistry Requirements'
 
                #    ws.cell(row=15+i*2,column=3).border = thin_border

                    
                #    ws.cell(row=15+i*2,column=8).value = "Chemistry is reported in Chemical sheet"
                #    ws.cell(row=15+i*2,column=8).border = thin_border

                #    continue

                #if(armdata["S_NO"][i].count('B2.1')>0):

                #    ws.cell(row=15+i*2,column=3).value = 'Additional Chemistry Requirements'
 
                #    ws.cell(row=15+i*2,column=3).border = thin_border

                #    ws.cell(row=15+i*2,column=8).value = "Chemistry is reported in Chemical sheet"
                #    ws.cell(row=15+i*2,column=8).border = thin_border

                #    continue
         


                ws.cell(row=15+i*2,column=1).value = armdata["ROW_INDEX"][i]
                ws.cell(row=15+i*2,column=1).border = thin_border


                ws.cell(row=15+i*2,column=2).value = armdata["S_NO"][i]
                ws.cell(row=15+i*2,column=2).border = thin_border

            
                ws.cell(row=15+i*2,column=3).value = armdata["Parameter"][i]
                ws.cell(row=15+i*2,column=3).border = thin_border

                ws.cell(row=15+i*2,column=4).value = "ARM"
                ws.cell(row=15+i*2,column=4).border = thin_border
                ws.cell(row=15+i*2+1,column=4).value = "ASME"
                ws.cell(row=15+i*2+1,column=4).border = thin_border

                ws.cell(row=15+i*2,column=5).value = armdata["Reference_Code_Value"][i] 
                ws.cell(row=15+i*2,column=5).border = thin_border
            
                if(armdata["Parameter"][i].count('ARM') > 0 or armdata["Parameter"][i].lower().count('rev no') > 0):

                    ws.cell(row=15+i*2,column=6).value = ARMNo
                    ws.cell(row=15+i*2,column=6).border = thin_border
                    if(ARMNo.count(ws.cell(row=15+i*2,column=5).value)>0):
                        ws.cell(row=15+i*2,column=7).value = "Accepted"
                        ws.cell(row=15+i*2,column=7).border = thin_border
                        ws.cell(row=15+i*2,column=7).fill = greenFill
                    else:
                        Total_deviation.add('ARM')
                        ws.cell(row=15+i*2,column=7).value = "Rejected"
                        ws.cell(row=15+i*2,column=7).fill = redFill
            
                elif(armdata["Parameter"][i].count('Material Specification') > 0):

                    ws.cell(row=15+i*2,column=6).value = Material1
                    ws.cell(row=15+i*2,column=6).border = thin_border
                    if(Material1.lower().count(ws.cell(row=15+i*2,column=5).value.lower().split(' ')[0].replace('-','')) >0):
                        ws.cell(row=15+i*2,column=7).value = "Accepted"
                        ws.cell(row=15+i*2,column=7).border = thin_border
                        ws.cell(row=15+i*2,column=7).fill = greenFill
                    else:
                        ws.cell(row=15+i*2,column=7).value = "Not Compared"

                elif(armdata["Parameter"][i].count('Delivery') > 0):
                    print(Product)
                    conditions = ws.cell(row=15+i*2,column=5).value
                       
                    
                    if(Product!=None ):

                        ws.cell(row=15+i*2,column=6).value = Product

                        if(not np.isnan(Delivery)):
                            ws.cell(row=15+i*2,column=6).value = ws.cell(row=15+i*2,column=6).value+ ' , ' + Delivery
                            

                        flag = False
                        for k in conditions.split(' ') :
                            if ( ws.cell(row=15+i*2,column=6).value.lower().count(k.lower()) > 0 ):
                                flag=True
                        if(flag):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            ws.cell(row=15+i*2,column=7).fill = greenFill
                        else:
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            ws.cell(row=15+i*2,column=7).fill = redFill
                        #if('Delivery' in list_product):
                        #    ws.cell(row=15+i*2,column=8).value = "Delivery Condition is reported in product sheet"
                        #    ws.cell(row=15+i*2,column=8).border = thin_border
                    else:
                        ws.cell(row=15+i*2,column=7).value = "Not Compared"
                        ws.cell(row=15+i*2,column=7).border = thin_border

                elif(armdata["Parameter"][i].count('Steel making') > 0):

                    print('Process: ' + Process)
                    if(Process!=None):
                        ws.cell(row=15+i*2,column=6).value = Process
                        ws.cell(row=15+i*2,column=6).border = thin_border
                    
                    ws.cell(row=15+i*2,column=7).value = "Not Compared"


                    #if(certification.count(ws.cell(row=15+i*2,column=5).value)>0):
                    #    ws.cell(row=15+i*2,column=7).value = "Accepted"
                    #    ws.cell(row=15+i*2,column=7).fill = greenFill
                    #else:
                    #    Total_deviation.add('Certification')
                    #    ws.cell(row=15+i*2,column=7).value = "Rejected"
                    #    ws.cell(row=15+i*2,column=7).fill = redFill
                    #    ws.cell(row=15+i*2,column=7).border = thin_border

                elif(armdata["Parameter"][i].count('Certification') > 0):

                    ws.cell(row=15+i*2,column=6).value = certification
                    ws.cell(row=15+i*2,column=6).border = thin_border
                    print(certification)

                    if(certification.count(ws.cell(row=15+i*2,column=5).value.split(' ')[-1])>0):
                        ws.cell(row=15+i*2,column=7).value = "Accepted"
                        ws.cell(row=15+i*2,column=7).fill = greenFill
                    else:
                        Total_deviation.add('Certification')
                        ws.cell(row=15+i*2,column=7).value = "Rejected"
                        ws.cell(row=15+i*2,column=7).fill = redFill
                        ws.cell(row=15+i*2,column=7).border = thin_border

                #Inspection By

                elif(armdata["Parameter"][i].count('Inspection By') > 0):
                    if(tpi!=None):

                        ws.cell(row=15+i*2,column=6).value = tpi
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        print(tpi)
                        ws.cell(row=15+i*2,column=7).value = "Not Compared"

                elif(armdata["Parameter"][i].count('CHEMISTRY') > 0):

                    ws.cell(row=15+i*2,column=8).value = "Chemistry is reported in Chemical sheet"
                    ws.cell(row=15+i*2,column=8).border = thin_border
                    
                    print("Chemical deviation - " + str(Chemical_deviation))
                    if(len(Chemical_deviation)>0):
                        Total_deviation.add(str(Chemical_deviation))
                        ws.cell(row=15+i*2,column=7).value = "Rejected"
                        ws.cell(row=15+i*2,column=7).fill = redFill
                    else:
                        ws.cell(row=15+i*2,column=7).value = "Accepted"
                        ws.cell(row=15+i*2,column=7).fill = greenFill 
                        ws.cell(row=15+i*2,column=7).border = thin_border

                elif(armdata["Parameter"][i].lower().count('product analysis') > 0):
                    if(d['Product Analysis']!=None):
                        ws.cell(row=15+i*2,column=6).value = d['Product Analysis']
                        ws.cell(row=15+i*2,column=6).border = thin_border  
                    
                    ws.cell(row=15+i*2,column=7).value = "Not Compared"
                    
                    ws.cell(row=15+i*2,column=8).value = "Chemistry is reported in Chemical sheet"
                    ws.cell(row=15+i*2,column=8).border = thin_border

                elif(armdata["Parameter"][i].count('MECHANICAL TESTING') > 0):

                    ws.cell(row=15+i*2,column=8).value = "Mechanical is reported in Mechanical sheet"
                    ws.cell(row=15+i*2,column=8).border = thin_border
                    print("Mechanical Deviation - " + str(Mechanical_deviation))

                    if(len(Mechanical_deviation)>0):
                        Total_deviation.add(str(Mechanical_deviation))
                        ws.cell(row=15+i*2,column=7).value = "Rejected"
                        ws.cell(row=15+i*2,column=7).fill = redFill
                    else:
                        ws.cell(row=15+i*2,column=7).value = "Accepted"
                        ws.cell(row=15+i*2,column=7).fill = greenFill 
                        ws.cell(row=15+i*2,column=7).border = thin_border
                        #Hardness
                elif(armdata["Parameter"][i].count('Hardness') > 0):

                    ws.cell(row=15+i*2,column=8).value = "Hardness is reported in Mechanical sheet"
                    ws.cell(row=15+i*2,column=8).border = thin_border

                    if('Hardness' in Mechanical_deviation):

                        ws.cell(row=15+i*2,column=7).value = "Rejected"
                        ws.cell(row=15+i*2,column=7).fill = redFill
                    else:
                        ws.cell(row=15+i*2,column=7).value = "Accepted"
                        ws.cell(row=15+i*2,column=7).fill = greenFill 
                        ws.cell(row=15+i*2,column=7).border = thin_border

                elif(armdata["Parameter"][i].lower().count('inter-granular corrosion') > 0):

                    if(d['IGC']!=None):


                        ws.cell(row=15+i*2,column=6).value = d['IGC']
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        
                        if(d['IGC'].lower().count('igc') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            if(np.isnan(list_product['IGC'][0])):
                                pass
                            else:
                                ws.cell(row=15+i*2,column=8).value = "IGC is reported in product description sheet."
                        else:
                            Total_deviation.add('IGC')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                        print(armdata["Parameter"][i+1], armdata["Parameter"][i+2])

                        if(armdata["Parameter"][i+1].lower().count('procedure') >0 or armdata["Parameter"][i+1].lower().count('acceptance') >0):
                            Acceptance(ws,'igc',d['IGC'].lower(),i+1)

                        if(armdata["Parameter"][i+2].lower().count('acceptance') >0 or armdata["Parameter"][i+2].lower().count('procedure') >0):
                            Acceptance(ws,'igc',d['IGC'].lower(),i+2)
                    else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"

                elif(armdata["Parameter"][i].lower().count('eddy current') > 0):
                    

                   if(d['Eddy Current']!=None):

                        ws.cell(row=15+i*2,column=6).value = d['Eddy Current']
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        
                        if(d['Eddy Current'].lower().count('eddy current') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            if(not np.isnan(list_product['ECT'][0])):
                                ws.cell(row=15+i*2,column=8).value = "ECT is reported in product description sheet."
                        else:
                            Total_deviation.add('ECT')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                        print(armdata["Parameter"][i+1], armdata["Parameter"][i+2])

                        if(armdata["Parameter"][i+1].lower().count('procedure') >0 or armdata["Parameter"][i+1].lower().count('acceptance') >0):
                            Acceptance(ws,'eddy current',d['Eddy Current'].lower(),i+1)

                        if(armdata["Parameter"][i+2].lower().count('acceptance') >0 or armdata["Parameter"][i+2].lower().count('procedure') >0):
                            Acceptance(ws,'eddy current',d['Eddy Current'].lower(),i+2)
                   else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"

                elif(armdata["Parameter"][i].lower().count('hydrostatic testing') > 0 or armdata["Parameter"][i].upper().count('HYDRO STATIC TEST') > 0):

                    if(d['Hydrostatic Test']!=None):

                        ws.cell(row=15+i*2,column=6).value = d['Hydrostatic Test']
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        
                        if(d['Hydrostatic Test'].lower().count('hydrostatic test') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            if(not np.isnan(list_product['HT'][0])):
                                ws.cell(row=15+i*2,column=8).value = "HT is reported in product description sheet."

                        else:
                            Total_deviation.add('HT')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                        print(armdata["Parameter"][i+1], armdata["Parameter"][i+2])

                        if(armdata["Parameter"][i+1].lower().count('procedure') >0 or armdata["Parameter"][i+1].lower().count('acceptance') >0):
                            Acceptance(ws,'hydrostatic test',d['Hydrostatic Test'].lower(),i+1)

                        if(armdata["Parameter"][i+2].lower().count('acceptance') >0 or armdata["Parameter"][i+2].lower().count('procedure') >0):
                            Acceptance(ws,'hydrostatic test',d['Hydrostatic Test'].lower(),i+2)
                    else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"
                elif(armdata["Parameter"][i].lower().count('positive material identification') > 0):
                    if(d['PMI']!=None):

                        ws.cell(row=15+i*2,column=6).value = d['PMI']
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        
                        if(d['PMI'].lower().count('pmi') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            if(not np.isnan(list_product['PMI'][0])):
                                ws.cell(row=15+i*2,column=8).value = "PMI is reported in product description sheet."

                        else:
                            Total_deviation.add('PMI')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                        print(armdata["Parameter"][i+1], armdata["Parameter"][i+2])

                        if(armdata["Parameter"][i+1].lower().count('procedure') >0 or armdata["Parameter"][i+1].lower().count('acceptance') >0):
                            Acceptance(ws,'pmi',d['PMI'].lower(),i+1)

                        if(armdata["Parameter"][i+2].lower().count('acceptance') >0 or armdata["Parameter"][i+2].lower().count('procedure') >0):
                            Acceptance(ws,'pmi',d['PMI'].lower(),i+2)
                    else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"
                elif(armdata["Parameter"][i].count('Repair Welding') > 0):

                    if(d['Weld']!=None):

                        ws.cell(row=15+i*2,column=6).value = d['Weld']
                        ws.cell(row=15+i*2,column=6).border = thin_border

                        if(d['Weld'].lower().count('no') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                        else:
                            Total_deviation.add('Weld repair')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                    else:

                        ws.cell(row=15+i*2,column=7).value = "Not Compared"
                elif(armdata["Parameter"][i].count('Visual Inspection') > 0):


                    if(d['Visual']!=None):

                        ws.cell(row=15+i*2,column=6).value = d['Visual']
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        
                        if(d['Visual'].lower().count('visual') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                        else:
                            Total_deviation.add('Visual')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                    else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"

                elif(armdata["Parameter"][i].count('PT examination') > 0):


                    if(d['PT examination']!=None):

                        ws.cell(row=15+i*2,column=6).value = d['PT examination']
                        ws.cell(row=15+i*2,column=6).border = thin_border
                        
                        if(d['PT examination'].lower().count('pt examination') >0):

                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                            if(not np.isnan(list_product['PT'][0])):
                                ws.cell(row=15+i*2,column=8).value = "PT is reported in product description sheet."
                        else:
                            Total_deviation.add('PT examination')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                        print(armdata["Parameter"][i+1], armdata["Parameter"][i+2])

                        if(armdata["Parameter"][i+1].lower().count('procedure') >0 or armdata["Parameter"][i+1].lower().count('acceptance') >0):
                            Acceptance(ws,'pt examination',d['PT examination'].lower(),i+1)

                        if(armdata["Parameter"][i+2].lower().count('acceptance') >0 or armdata["Parameter"][i+2].lower().count('procedure') >0):
                            Acceptance(ws,'pt examination',d['PT examination'].lower(),i+2)
                    else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"

                elif(armdata["S_NO"][i].count('D1') > 0):
                    if(ContentHeatTreatment!=None):


                        ws.cell(row=15+i*2,column=6).value = ContentHeatTreatment

                        ws.cell(row=15+i*2,column=6).border = thin_border

                        ws.cell(row=15+i*2+1,column=5).value = stHeatTreatment
                        ws.cell(row=15+i*2+1,column=5).border = thin_border
                        print(ContentHeatTreatment,stHeatTreatment)

                        if(ContentHeatTreatment.lower().count(stHeatTreatment.split()[0].lower()) > 0):
                            ws.cell(row=15+i*2,column=7).value = "Accepted"
                            ws.cell(row=15+i*2,column=7).fill = greenFill 
                            ws.cell(row=15+i*2,column=7).border = thin_border
                        else:
                            Total_deviation.add('Heat Treatment')
                            ws.cell(row=15+i*2,column=7).value = "Rejected"
                            ws.cell(row=15+i*2,column=7).fill = redFill
                    else:
                       ws.cell(row=15+i*2,column=7).value = "Not Compared"

                
              
            except:
                print("error in arm data extraction")
                print("Oops!", sys.exc_info()[0], "occurred.")
                pass
            #print(armdata.iloc[i])
        ws.cell(row=15+i*2+1,column=3).value = "Remarks"
        for j in Additonal_remarks:
            if(len(j) > 25):
                ws.cell(row=15+i*2+1,column=5).value = j
                i+=1 
        print("ARM reporting completed")
        return Total_deviation

#ASME Section extraction
def getAsmeData(wb,CertificationDetailsPath, Material,Grade,TC_NO):
    print("Chemical started ..")
    # open ASME data sheet
    asmeData = pd.read_excel(path+ASMEFilename,header=2)
    # load chemical worksheet
    ws = wb.worksheets[1]

    mechanical = wb.worksheets[2]
    Chemical_deviation = set()
    Mechanical_deviation = set()
    product = wb.worksheets[3]
    HeatTreatment =''
    
    mechanical.cell(row=1,column = 1).value = "mechanical"
    
    product.cell(row=1,column =1).value = "Product"

    #print("asme data : ",asmeData)
    #try:
    #    # take grade 
    #print([i for i in asmeData.keys()])

    d ={}
    m = {}
    # Chemical from ASME
    for i in range(len(asmeData["Material"])):
        #if(asmeData["Material"][i]== Material and asmeData["Grade "][i]==Grade):
        if(Material.count(str(asmeData["Material"][i]))>0 and Grade.count(asmeData["Grade "][i])>0):
            # chemical
            d["C"] = { "min" : asmeData.iloc[i]["Carbon (%)"] , "max" : asmeData.iloc[i]["Unnamed: 7"] }
            d["Mn"] = { "min" :  asmeData.iloc[i]["Manganese (%)"], "max" : asmeData.iloc[i]["Unnamed: 21"]}
            d["P"] = { "min" :  asmeData.iloc[i]["Phosphorus (%)"], "max" : asmeData.iloc[i]["Unnamed: "+str(21+14)]}
            d["S"] = { "min" :  asmeData.iloc[i]["Sulfur (%)"], "max" : asmeData.iloc[i]["Unnamed: "+str(21+14+6)]}
            d["Si"] = { "min" :   asmeData.iloc[i]["Silicon (%)"], "max" : asmeData.iloc[i]["Unnamed: "+str(21+14+6+6)]}
            d["Ni"] = { "min" :  asmeData.iloc[i]["Nickel (%)"], "max" : asmeData.iloc[i]["Unnamed: "+str(21+14+6+6+6)]}
        
            d["Cr"] = { "min" :  asmeData.iloc[i]["Chromium (%)"], "max" : asmeData.iloc[i]["Unnamed: "+ str(21+14+6+6+6+6)] }
        
            d["Mo"] = { "min" :  asmeData.iloc[i]["Moly (%)"], "max" : asmeData.iloc[i]["Unnamed: "+ str(21+14+6+6+6+6+6)]}
            d["N2"] = { "min" :  asmeData.iloc[i]["Nitrogen (%)"], "max" : asmeData.iloc[i]["Unnamed: "+ str(21+14+6+6+6+6+6+6)]}
            d["Ti"] = { "min" :  asmeData.iloc[i]["Titanium  (%)"], "max" : asmeData.iloc[i]["Unnamed: "+ str(21+14+6+6+6+6+6+6+8)]}
            d["Cu"] = { "min" :  asmeData.iloc[i]["Copper (%)"], "max" : asmeData.iloc[i]["Unnamed: "+ str(21+14+6+6+6+6+6+6+2)]}
            
     
            # mechanical 
            print(asmeData.iloc[i]["""Tensile
Strength, ksi
[MPa]"""])
            m["TS"] = asmeData.iloc[i]["""Tensile
Strength, ksi
[MPa]"""]
            print(asmeData.iloc[i]["""Yield
Strength,
min, ksi
[MPa]"""])
            m["YS"] = asmeData.iloc[i]["""Yield
Strength,
min, ksi
[MPa]"""]
            print(asmeData.iloc[i]["Elongation in 2 in. or 50 mm, min, %B,C"])
            m["Elong"] = asmeData.iloc[i]["Elongation in 2 in. or 50 mm, min, %B,C"]
            print(asmeData.iloc[i]['Unnamed: 165'])

            print(asmeData.iloc[i]['Unnamed: 166'])
            m["Hardness"] = asmeData.iloc[i]['Unnamed: 166']

            HeatTreatment = asmeData.iloc[i]['Heat Treat Type']
            print(HeatTreatment)
            #print(d)
#        #print(Material.count(str(asmeData["Material"][i]))>0)k
    # compare data and write it to final report

    ChemicalTable_fileName = CertificationDetailsPath +  '\ChemicalTable.xlsx'
    ChemicalTable1_fileName = CertificationDetailsPath + '\ChemicalTable1.xlsx'
    #MechanicalTable_fileName = CertificationDetailsPath + '\MechanicalTable.xlsx'

    ProductTable_fileName = CertificationDetailsPath + '\ProdutDetails.xlsx'
    ProductTable1_fileName = CertificationDetailsPath + '\ProdutDetails2.xlsx'

    mtcChemicalData = pd.read_excel(ChemicalTable_fileName)
    mtcChemicalData1 = pd.read_excel(ChemicalTable1_fileName)
    del mtcChemicalData1['MTC1_ROW_INDEX']
    ##mtcMechanicalData = pd.read_excel(MechanicalTable_fileName)
    mtcChemicalData = pd.concat([mtcChemicalData,mtcChemicalData1],axis=1)

    mtcProductData = pd.read_excel(ProductTable_fileName)
    mtcProductData1 = pd.read_excel(ProductTable1_fileName)
    del mtcProductData1['MTC1_ROW_INDEX']
    mtcProductData = pd.concat([mtcProductData,mtcProductData1[['FLT','DET','PMI','ECT','VDI','IGC','ENDFINISH','Delivery']]],axis=1)

    #list_chemical = mtcChemicalData.keys()
    #print(mtcChemicalData.keys())
    list_chemical = mtcChemicalData.keys()
    #list_chemical = ['C',"Mn","P","S","Si","Cr","Ni","Mo","N2","Ti"]
    print(list_chemical)
    
   
    list_product = mtcProductData.keys()

    print(list_product)


    for i in range(1,len(list_chemical)):
        
        ws.cell(row=7,column=2+i).value = list_chemical[i]
        
        
        for j in range(len(mtcChemicalData['C'])):   
            try:
                if(list_chemical[i] =="Analysis"):
                    ws.cell(row=11+j,column=2+i).value = mtcChemicalData[list_chemical[i]][j]
                    continue

                if(i<3):
                    ws.cell(row=11+j,column=2+i).value = mtcChemicalData[list_chemical[i]][j]
                if(i>2):
                    print(list_chemical[i] , d[list_chemical[i]],mtcChemicalData[list_chemical[i]][j])
                    ws.cell(row=11+j,column=2+i).value = mtcChemicalData[list_chemical[i]][j]
                    ws.cell(row=8,column = 2 + i).value = d[list_chemical[i]]['min']
                    ws.cell(row=9,column = 2 + i).value = d[list_chemical[i]]['max'] 
                    
                    if(np.isnan(ws.cell(row=11+j,column=2+i).value)):
                        continue
                    else:
                        #print(ws.cell(row=11+j,column=2+i).value,ws.cell(row=8,column = 2 + i).value,ws.cell(row=9,column = 2 + i).value)
                        minCheck = True
                        maxCheck = True
                        if(d[list_chemical[i]]['min'] == '...' and d[list_chemical[i]]['max'] == '...'):
                            minCheck = True
                            maxCheck=True
                        elif(d[list_chemical[i]]['max'] == '...' and d[list_chemical[i]]['min'] != '...'):
                            if(float(mtcChemicalData[list_chemical[i]][j]) < float(d[list_chemical[i]]['min'] ) ):
                                minCheck = False
                     
                        elif(d[list_chemical[i]]['max'] != '...' and d[list_chemical[i]]['min'] == '...'):
                            if(float(mtcChemicalData[list_chemical[i]][j]) > float(d[list_chemical[i]]['max'] ) ):
                                maxCheck = False
                   
                        else:

                            if(float(mtcChemicalData[list_chemical[i]][j]) > float(d[list_chemical[i]]['max'] ) ):
                                maxCheck = False
                            if(float(mtcChemicalData[list_chemical[i]][j]) < float(d[list_chemical[i]]['min'] ) ):
                                minCheck = False
                        if(maxCheck and minCheck):
                            ws.cell(row=11+j,column=2+i).fill = greenFill 
                        else:
                            Chemical_deviation.add(list_chemical[i])
                            ws.cell(row=11+j,column=2+i).fill = redFill 
            except:
                print('error chemical in comparison')
                pass


            #if(i>2 and ws.cell(row=8,column = 2 + i).value != '...' and int(ws.cell(row=8,column = 2 + i).value) > mtcChemicalData[list_chemical[i]][j] )
            #ws.cell(row=10+j,column=2+i).fill = my_fill
    if(len(Chemical_deviation)>0):
        ws.cell(row=5,column=2).value = 'Deviation' + str(Chemical_deviation)
        ws.cell(row=5,column=2).fill = redFill
    else:
        ws.cell(row=5,column=2).value = 'Accepted'
        ws.cell(row=5,column=2).fill = greenFill
    print("Chemical Completed")

    MechanicalTable_fileName = CertificationDetailsPath + '\MechanicalTable.xlsx'

    mtcMechanicalData = pd.read_excel(MechanicalTable_fileName)

    print(mtcMechanicalData)
    
    list_mechanical = mtcMechanicalData.keys()

    print(list_mechanical)
    Delivery = ''
    for i in range(1,len(list_mechanical)):
        mechanical.cell(row=7,column=1+i).value = list_mechanical[i]
        mechanical.cell(row=7,column=1+i).border = thin_border
        

        for j in range(len(mtcMechanicalData[list_mechanical[i]])):
            #print(mtcMechanicalData[list_mechanical[i]][j])
            try:
                if(list_chemical[i].count('Delivery')):
                    Delivery = Delivery + ' ' + mtcMechanicalData[list_mechanical[i]][j]
                mechanical.cell(row=10+j,column=1+i).value = mtcMechanicalData[list_mechanical[i]][j]
                mechanical.cell(row=10+j,column=1+i).border = thin_border
                if(i>2 and i < 5):
                    mechanical.cell(row=7,column=i).value = list_mechanical[i-1] + "( N/mm2 )"
                    mechanical.cell(row=7,column=i).border = thin_border
                    mechanical.cell(row=8,column=i).value = m[list_mechanical[i-1]]
                    if(np.isnan(mechanical.cell(row=10+j,column=i).value)):
                        continue
                    else:
                        print(type(mechanical.cell(row=10+j,column=i).value), mechanical.cell(row=8,column=i).value.split(' ')[1][1:-1])
                        if(int(mechanical.cell(row=10+j,column=i).value) > int(mechanical.cell(row=8,column=i).value.split(' ')[1][1:-1])):
                            mechanical.cell(row=10+j,column=i).fill = greenFill
                        else:
                            Mechanical_deviation.add(mechanical.cell(row=7,column=i).value)
                            mechanical.cell(row=10+j,column=i).fill = redFill
                    
                if(i>4 and i < 6):

                    mechanical.cell(row=8,column=i).value = m[list_mechanical[i-1]]
                    if(np.isnan(mechanical.cell(row=10+j,column=i).value)):
                        continue
                    else:
                        
                        print(type(mechanical.cell(row=10+j,column=i).value), mechanical.cell(row=8,column=i).value.split(' ')[0].split('-')[0])

                        if(int(mechanical.cell(row=10+j,column=i).value) > int(mechanical.cell(row=8,column=i).value.split(' ')[0].split('-')[0])):
                            mechanical.cell(row=10+j,column=i).fill = greenFill
                        else:
                            Mechanical_deviation.add(mechanical.cell(row=7,column=i).value)
                            mechanical.cell(row=10+j,column=i).fill = redFill
                
                if(i>5 and i < 7):
                    mechanical.cell(row=8,column=i).value = m[list_mechanical[i-1]]
                    print(mechanical.cell(row=10+j,column=i).value.split('-')[0], mechanical.cell(row=8,column=i).value.split(' ')[0].split('-')[0])
                                            
                    if(int(mechanical.cell(row=10+j,column=i).value.split('-')[0]) < int(mechanical.cell(row=8,column=i).value.split(' ')[0].split('-')[0])):
                        mechanical.cell(row=10+j,column=i).fill = greenFill
                    else:
                        Mechanical_deviation.add(mechanical.cell(row=7,column=i).value)
                        mechanical.cell(row=10+j,column=i).fill = redFill
            except:
                    print('error in mechanical comparison')
                    print("Oops!", sys.exc_info()[0], "occurred.")
                    pass

    if(len(Mechanical_deviation)>0):
        mechanical.cell(row=5,column=2).value = 'Deviation' + str(Mechanical_deviation)
        mechanical.cell(row=5,column=2).fill = redFill
    else:
        mechanical.cell(row=5,column=2).value = 'Accepted'
        mechanical.cell(row=5,column=2).fill = greenFill
    
    print("Mechanical Completed")

    
    for i in range(1,len(list_product)):
        product.cell(row=7,column=i).value = list_product[i]
        product.cell(row=7,column=i).border = thin_border
        for j in range(len(mtcProductData[list_product[i]])):
            #print(mtcProductData[list_product[i]][j])
            try:
                product.cell(row=8+j,column=i).value = mtcProductData[list_product[i]][j]
                product.cell(row=8+j,column=i).border = thin_border
            except:
                    print('error in product comparison')
                    
                    print("Oops!", sys.exc_info()[0], "occurred.")
                    pass
    print("Product Completed")
    return Chemical_deviation,Mechanical_deviation,HeatTreatment,Delivery,mtcProductData
    #except:
    #    print("Error while ASME data comparison")
def document_select_function(name):
    Tc_id_query = """ SELECT TOP 1 [name]
      ,[url]
      ,[status]
      ,[excel_url]
      ,[type]
      ,[id]
  FROM [dbo].[instoman1_document]
  where name like '%{}%'
  ORDER BY ID DESC """.format( name.split('_')[0])
    print(Tc_id_query)
    cursor.execute(Tc_id_query)
    try:

        id = cursor.fetchone()[5]
        print(id)
    except:
        id = 0
    return id

def StartComparison(CertificationDetailsPath):
    #Certification_fileName = CertificationDetailsPath.split('//')[-1]+'.xlsx'
    wb = openpyxl.load_workbook(path+ReportFilename,read_only=False, keep_vba=True)
    
    #      i.split('\\')[-1]
    tc_id = document_select_function(CertificationDetailsPath.split('\\')[-1])
    
    TC_details = pd.read_excel(CertificationDetailsPath + ".xlsx")

    try:
        Total_deviation = set()

        print(TC_details.keys())
 
        Material = TC_details["Specification"]
        


        Material1 = Material[0].replace(' ','')

        print('Specification: ' + Material1)


        print(Material1)

        Grade = TC_details["Grade"][0]
        
        #Grade = Grade[0].split(':')[1].split('(')[0].replace(' ','')
        
        print(Grade)
        

        TC_NO = TC_details["TcNo"][0]
  
        print(TC_NO)
        TC_Date = TC_details["TcDate"][0]
        

        PO_NO  = TC_details["PoNo"][0]


        PO_Date = TC_details["PODate"][0]
        Process = TC_details["Process"][0]
        Product = TC_details["Product"][0]
        certification  = TC_details["Certification"][0]
        
        print(Product)

        ARMNo = TC_details["ARMNo"][0].split(':')[1] if len(TC_details["ARMNo"][0].split(':')) > 1 else ''
        #REV = TC_details['Rev'][0].split(':')[1] if len(TC_details['Rev'][0].split(':')) > 1 else 'None'
        try:
            Rev1 = TC_details["Rev1"][0]
            Rev = TC_details["Rev"][0].split(' ')[0]

            if(ARMNo==''):
                ARMNo =  Rev+ ' '  + Rev1
            else:
                print(ARMNo)
                ARMNo = ARMNo + ' ' + Rev1
        except:
            pass
        print(Material[0],Grade,TC_NO,TC_Date,PO_NO,PO_Date,ARMNo)
        
        insert_query1 = """ INSERT INTO [dbo].[instoman1_admins]
           ([tc_name]
           ,[arm_name]
           ,[material_code]
           ,[vendor_name]
           ,[tcs_id_id])
     VALUES
           ('{}'
           ,'{}'
           ,'{}'
           ,'Chandan Steel'
           ,{} ) """.format(TC_NO, ARMNo, Material[0], tc_id)
        print(insert_query1)
        cursor.execute(insert_query1)
        

        cursor.commit()
        
        cursor.execute("SELECT @@IDENTITY AS ID;")
        current_id = int(cursor.fetchone()[0])

        ws = wb.worksheets[0]
        ws.cell(row=5,column=2).value = TC_NO
        ws.cell(row=5,column=6).value = ARMNo
        ws.cell(row=6,column=2).value = TC_Date
        ws.cell(row=7,column=2).value = PO_NO
        ws.cell(row=7,column=6).value = 'Chadan Steel Limited'
    
    
        ws.cell(row=8,column=2).value = Product

        ws.cell(row=6,column=2).value = TC_Date
        ws.cell(row=6,column=6).value = Material[0]
        ws.cell(row=8,column=6).value = Grade
        print(ws.cell(row=6,column=6).value)

        HeatTreatment = ''
        Delivery = ''

        Chemical_deviation,Mechanical_deviation,HeatTreatment,Delivery,mtcProductData = getAsmeData(wb,CertificationDetailsPath,Material1,Grade.replace(' ',''),TC_NO)
     
        if(Delivery==''):
            Delivery = TC_details["Delivery"][0]
        print("Delivery " , Delivery)
        print('wait for arm .. ',ARMNo)
        count = 4
        while(count>0):
            for i in armDataset:
                if(i['name'].count(str(ARMNo.split()[0]))>0):
                        arm_id = document_select_function(ARMNo.split()[0])
                        Total_deviation = write_arm(wb,CertificationDetailsPath,i['url'],ARMNo,Material1,Grade,Chemical_deviation,Mechanical_deviation,HeatTreatment,certification,Process,Product,Delivery,mtcProductData)
                        i["arm"] = str(ARMNo.split()[0])
                        
                        count=-1
                        break

            print('can not find arm')
            time.sleep(5)
            count-=1
        if(count<1):
            ws.cell(row=10,column=1).value = "No ARM found!"
            arm_id = 0
        print(Total_deviation)

        if(len(Total_deviation) > 0):
            ws.cell(row=9,column=2).value =  "TC is Not acceptable -"+ str(Total_deviation)
            ws.cell(row=9,column=2).fill = redFill
        else:
            if(len(Mechanical_deviation) > 0 or len(Chemical_deviation)>0):
                ws.cell(row=9,column=2).value =  "TC is Not acceptable -" + str(Mechanical_deviation) + str(Chemical_deviation) 
                ws.cell(row=9,column=2).fill = redFill
            else:
                ws.cell(row=9,column=2).value =  "TC is acceptable"
                ws.cell(row=9,column=2).fill = greenFill


        report_path = path + 'final_report/' + TC_NO.replace('/','-').replace('.','').replace(' ','').replace(':','')+ '.xlsm'
        #    write_productDetails(CertificationDetailsPath,TC_No)

        #search ARM id from document table and get id.
        #enter tc details into the database tc, arm , tc_id, arm_id, material+grade, path
        #update report path of the id


        update_query = """ UPDATE [dbo].[instoman1_admins]
   SET 
      [report_path] = '{}' 
      ,[arm_id] = {}
 WHERE [id] like {} """.format(report_path,arm_id,current_id)
        print(update_query)
        cursor.execute(update_query)

        cursor.commit()

        wb.save(report_path)
        
        
        print(TC_NO + ': inspection completed ')
    except:
        print('error while tc comparison')
        print("Oops!", sys.exc_info()[0], "occurred.")
        pass

    #except:
    #    print("Error in TC data extraction")
dataset = []

def document_update_function(i,type):
    query = """ UPDATE [dbo].[instoman1_document]
                            SET [status] = 'true',
                                [excel_url] = '{}'
                                ,[type] = '{}'
                            WHERE name like '%{}%'  """.format(i,type,i.split('\\')[-1])
    print(query)

    row = cursor.execute(query)
    conn.commit()
    print(row)
#StartComparison('//10.7.66.75/next 25/Chandan Tubes/MTC-Extraction/Export_MTC/Batch/Data_00000001')
def Main(path):
    time.sleep(5)
    li = glob.glob(path+"/*")
    if(len(li)>1):
        #print(li)
        for i in li:
            if(i.find(".tmp")  ==-1 and  i.find(".xlsx") ==-1 and i.find(".pwlf") ==-1):
                #StartComparison(i)
                document_update_function(i,'MTC')

                dataset.append({ 'name' : i.split('\\')[-1] , 'url': i , 'type' : 'mtc'})
                #enter tc excelsheet url into the database
                print(dataset)
                StartComparison(i)

armDataset =[]


def ARMMain(path):
    time.sleep(5)

    li = glob.glob(path+"/*")
    if(len(li)>1):
        #print(li)
        for i in li:
            if(i.find(".tmp")  ==-1 and  i.find(".xlsx") ==-1 and i.find(".pwlf") ==-1):
                #StartComparison(i)
                document_update_function(i,'ARM')
                armDataset.append({ 'name' : i.split('\\')[-1] , 'url': i , 'type' : 'arm'})
                #enter arm excel sheet url to database
                print(armDataset)
class MonitorFolder(FileSystemEventHandler):
    FILE_SIZE=1000
    
    def on_created(self, event):
         #print(event.src_path, event.event_type)
         #print(event.src_path.find(".tmp")  ==-1 and  event.src_path.find(".xlsx") ==-1 and event.src_path.find(".pwlf") ==-1)
         if(event.src_path.find(".tmp")  ==-1 and  event.src_path.find(".xlsx") ==-1 and event.src_path.find(".pwlf") ==-1  ):
             #print("root directory",event.src_path)
             Main(event.src_path)
             #StartComparison(event.src_path)
         #self.checkFolderSize(event.src_path)
   
    #def on_modified(self, event):
    #    print(event.src_path, event.event_type)
    #    #self.checkFolderSize(event.src_path)
    #def checkFolderSize(self,src_path):
    #    if os.path.isdir(src_path):
    #        if os.path.getsize(src_path) >self.FILE_SIZE:
    #            print("Time to backup the dir")
    #    else:
    #        if os.path.getsize(src_path) >self.FILE_SIZE:
                #print("very big file")
class MonitorArmFolder(FileSystemEventHandler):

    FILE_SIZE=1000
    def on_created(self, event):
         #print(event.src_path, event.event_type)
         #print(event.src_path.find(".tmp")  ==-1 and  event.src_path.find(".xlsx") ==-1 and event.src_path.find(".pwlf") ==-1)
         if(event.src_path.find(".tmp")  ==-1 and  event.src_path.find(".xlsx") ==-1 and event.src_path.find(".pwlf") ==-1  ):
             #print("root directory",event.src_path)
             ARMMain(event.src_path)

if __name__ == "__main__":
    src_path = r"\\10.7.66.75\e\next 25\Chandan Tubes\MTC-Extraction\Export_MTC"
    
    arm_path = r"\\10.7.66.75\e\next 25\Chandan Tubes\ARM-EXport"
    
    event_handler=MonitorFolder()

    
    observer = Observer()
    observer.schedule(event_handler, path=src_path, recursive=True)
    print("Monitoring started")
    observer.start()

    armEvent_handler = MonitorArmFolder()

    armobserver = Observer()
    armobserver.schedule(armEvent_handler, path=arm_path, recursive=True)
    print("Arm Monitoring started")
    armobserver.start()


    try:
        while(True):
           time.sleep(1)
           
    except KeyboardInterrupt:
            observer.stop()
            observer.join()
            armobserver.stop()
            armobserver.join()

 #save final format - 

#with zipfile.ZipFile(path, 'r') as z:
#    z.extractall('./xlsm/')

#with zipfile.ZipFile(r'\\10.23.67.120\c$\Users\90347908\Desktop\next 25\Chandan TUbes 2\Final_report\2.xlsx', 'r') as z:
#    z.extractall('./xlsx/')

#copyfile('./xlsm/[Content_Types].xml','./xlsx/[Content_Types].xml')
#copyfile('./xlsm/xl/_rels/workbook.xml.rels','./xlsx/xl/_rels/workbook.xml.rels')
#copyfile('./xlsm/xl/vbaProject.bin','./xlsx/xl/vbaProject.bin')

#z = zipfile.ZipFile(r'\\10.23.67.120\c$\Users\90347908\Desktop\next 25\Chandan TUbes 2\Final_report\2.zip', 'w')

#os.chdir('./xlsx')

#for root, dirs, files in os.walk('./'):
#        for file in files:
#            z.write(os.path.join(root, file))
#z.close()

#clean
#os.chdir(PAD)
#rmtree('./xlsm/')
#rmtree('./xlsx/')
#os.remove(r'\\10.23.67.120\c$\Users\90347908\Desktop\next 25\Chandan TUbes 2\Final_report\2.xlsx')
#os.rename(r'\\10.23.67.120\c$\Users\90347908\Desktop\next 25\Chandan TUbes 2\Final_report\2.zip', r'\\10.23.67.120\c$\Users\90347908\Desktop\next 25\Chandan TUbes 2\Final_report\2.xlsm')